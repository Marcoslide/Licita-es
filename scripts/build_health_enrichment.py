#!/usr/bin/env python3
"""Build the read-only BPS/CMED index used by the public procurement detail API."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sqlite3
import unicodedata
import zipfile
from datetime import UTC, datetime
from pathlib import Path


BPS_SOURCE_PAGE = "https://dadosabertos.saude.gov.br/dataset/bps"
CMED_SOURCE_PAGE = "https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos"


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char)).lower()
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def digits(value: object) -> str | None:
    result = re.sub(r"\D", "", str(value or ""))
    return result or None


def number(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    clean = re.sub(r"[^0-9,.-]", "", str(value)).strip()
    if not clean:
        return None
    if "," in clean:
        clean = clean.replace(".", "").replace(",", ".")
    try:
        return float(clean)
    except ValueError:
        return None


def date_iso(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    match = re.match(r"(\d{2})/(\d{2})/(\d{4})", text)
    if match:
        return f"{match[3]}-{match[2]}-{match[1]}"
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
    return f"{match[1]}-{match[2]}-{match[3]}" if match else None


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        pragma journal_mode = wal;
        pragma synchronous = normal;
        create table metadata (key text primary key, value text not null);
        create table bps_prices (
          id integer primary key,
          source_year integer not null,
          institution_name text,
          institution_cnpj text,
          institution_sphere text,
          city text,
          state_code text,
          purchase_date text,
          inserted_date text,
          modality text,
          purchase_type text,
          catmat_code text,
          description text not null,
          description_key text not null,
          unit text,
          unit_capacity text,
          generic_flag text,
          anvisa_registration text,
          supplier_cnpj text,
          supplier_name text,
          manufacturer_cnpj text,
          manufacturer_name text,
          quantity real,
          unit_price real,
          total_price real,
          source_url text not null
        );
        create index bps_catmat_idx on bps_prices(catmat_code, source_year desc);
        create index bps_anvisa_idx on bps_prices(anvisa_registration);
        create index bps_supplier_idx on bps_prices(supplier_cnpj, source_year desc);
        create index bps_state_idx on bps_prices(state_code, source_year desc);
        create virtual table bps_fts using fts5(
          description, content='bps_prices', content_rowid='id',
          tokenize='unicode61 remove_diacritics 2'
        );

        create table cmed_prices (
          id integer primary key,
          published_at text,
          substance text not null,
          substance_key text not null,
          company_cnpj text,
          laboratory text,
          ggrem_code text,
          registration text,
          ean text,
          product text,
          product_key text,
          presentation text,
          therapeutic_class text,
          product_type text,
          price_regime text,
          pf_no_tax real,
          pf_zero real,
          pf_eighteen real,
          pmvg_no_tax real,
          pmvg_zero real,
          pmvg_eighteen real,
          hospital_restriction text,
          cap text,
          confaz_87 text,
          tax_credit_list text,
          commercialization text,
          label text,
          source_url text not null
        );
        create index cmed_registration_idx on cmed_prices(registration);
        create index cmed_substance_idx on cmed_prices(substance_key);
        create index cmed_product_idx on cmed_prices(product_key);
        create virtual table cmed_fts using fts5(
          substance, product, presentation, content='cmed_prices', content_rowid='id',
          tokenize='unicode61 remove_diacritics 2'
        );
        """
    )


def import_bps(connection: sqlite3.Connection, paths: list[Path]) -> int:
    sql = """insert into bps_prices (
      source_year,institution_name,institution_cnpj,institution_sphere,city,state_code,
      purchase_date,inserted_date,modality,purchase_type,catmat_code,description,description_key,
      unit,unit_capacity,generic_flag,anvisa_registration,supplier_cnpj,supplier_name,
      manufacturer_cnpj,manufacturer_name,quantity,unit_price,total_price,source_url
    ) values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    total = 0
    batch: list[tuple[object, ...]] = []
    for path in paths:
        with zipfile.ZipFile(path) as archive:
            name = next(item for item in archive.namelist() if item.lower().endswith(".csv"))
            with archive.open(name) as binary:
                stream = io.TextIOWrapper(binary, encoding="utf-8-sig", errors="replace", newline="")
                for row in csv.DictReader(stream, delimiter=";"):
                    description = str(row.get("descricao_catmat") or "").strip()
                    if not description:
                        continue
                    year = int(number(row.get("ano_compra")) or Path(name).stem[:4] or 0)
                    batch.append((
                        year, row.get("nome_instituicao"), digits(row.get("cnpj_instituicao")),
                        row.get("esfera"), row.get("municipio_instituicao"), row.get("uf"),
                        date_iso(row.get("compra")), date_iso(row.get("insercao")),
                        row.get("modalidade_compra"), row.get("tipo_compra"), digits(row.get("codigo_br")),
                        description, normalize(description), row.get("unidade_fornecimento"),
                        row.get("unidade_fornecimento_capacidade"), row.get("generico"), digits(row.get("anvisa")),
                        digits(row.get("cnpj_fornecedor")), row.get("fornecedor"),
                        digits(row.get("cnpj_fabricante")), row.get("fabricante"),
                        number(row.get("qtd_itens_comprados")), number(row.get("preco_unitario")),
                        number(row.get("preco_total")), BPS_SOURCE_PAGE,
                    ))
                    if len(batch) >= 5000:
                        connection.executemany(sql, batch)
                        total += len(batch)
                        batch.clear()
        if batch:
            connection.executemany(sql, batch)
            total += len(batch)
            batch.clear()
        connection.commit()
    connection.execute("insert into bps_fts(rowid,description) select id,description from bps_prices")
    connection.commit()
    return total


def import_cmed(connection: sqlite3.Connection, path: Path) -> int:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - build-time dependency
        raise SystemExit("openpyxl is required only to build the CMED index") from exc
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    published_at = None
    header_row = None
    headers: list[str] = []
    for index, values in enumerate(sheet.iter_rows(min_row=1, max_row=80, values_only=True), 1):
        first = str(values[0] or "")
        if first.startswith("Publicada em"):
            published_at = first.replace("Publicada em", "").strip().rstrip(".")
        if normalize(first) == "substancia":
            header_row = index
            headers = [str(value or "").strip() for value in values]
            break
    if not header_row:
        raise SystemExit("CMED header row not found")
    sql = """insert into cmed_prices (
      published_at,substance,substance_key,company_cnpj,laboratory,ggrem_code,registration,ean,
      product,product_key,presentation,therapeutic_class,product_type,price_regime,
      pf_no_tax,pf_zero,pf_eighteen,pmvg_no_tax,pmvg_zero,pmvg_eighteen,
      hospital_restriction,cap,confaz_87,tax_credit_list,commercialization,label,source_url
    ) values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    total = 0
    batch: list[tuple[object, ...]] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = dict(zip(headers, values))
        substance = str(row.get("SUBSTÂNCIA") or "").strip()
        if not substance:
            continue
        product = str(row.get("PRODUTO") or "").strip()
        batch.append((
            published_at, substance, normalize(substance), digits(row.get("CNPJ")), row.get("LABORATÓRIO"),
            digits(row.get("CÓDIGO GGREM")), digits(row.get("REGISTRO")), digits(row.get("EAN 1")),
            product, normalize(product), row.get("APRESENTAÇÃO"), row.get("CLASSE TERAPÊUTICA"),
            row.get("TIPO DE PRODUTO (STATUS DO PRODUTO)"), row.get("REGIME DE PREÇO"),
            number(row.get("PF Sem Impostos")), number(row.get("PF 0%")), number(row.get("PF 18 %")),
            number(row.get("PMVG Sem Impostos")), number(row.get("PMVG 0 %")), number(row.get("PMVG 18 %")),
            row.get("RESTRIÇÃO HOSPITALAR"), row.get("CAP"), row.get("CONFAZ 87"),
            row.get("LISTA DE CONCESSÃO DE CRÉDITO TRIBUTÁRIO (PIS/COFINS)"),
            row.get("COMERCIALIZAÇÃO 2025"), row.get("TARJA"), CMED_SOURCE_PAGE,
        ))
        if len(batch) >= 3000:
            connection.executemany(sql, batch)
            total += len(batch)
            batch.clear()
    if batch:
        connection.executemany(sql, batch)
        total += len(batch)
    connection.execute(
        "insert into cmed_fts(rowid,substance,product,presentation) "
        "select id,substance,product,presentation from cmed_prices"
    )
    connection.commit()
    return total


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--bps", action="append", required=True, type=Path, help="Official BPS CSV ZIP; repeatable")
    parser.add_argument("--cmed", required=True, type=Path, help="Official CMED PMVG XLSX")
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.output.exists():
        args.output.unlink()
    connection = sqlite3.connect(args.output)
    try:
        create_schema(connection)
        bps_count = import_bps(connection, args.bps)
        cmed_count = import_cmed(connection, args.cmed)
        metadata = {
            "built_at": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
            "bps_records": str(bps_count),
            "cmed_records": str(cmed_count),
            "bps_sources": json.dumps([path.name for path in args.bps]),
            "cmed_source": args.cmed.name,
        }
        connection.executemany("insert into metadata(key,value) values (?,?)", metadata.items())
        connection.execute("analyze")
        connection.commit()
    finally:
        connection.close()
    digest = hashlib.sha256(args.output.read_bytes()).hexdigest()
    print(json.dumps({"output": str(args.output), "bps": bps_count, "cmed": cmed_count, "sha256": digest}))


if __name__ == "__main__":
    main()
